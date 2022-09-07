import os
import pandas as pd
import openpyxl
import numpy as np
from openpyxl import load_workbook
import interface_games_afs as interface

filename = "{}\jogos_excel.xlsx".format(os.getcwd())
print(filename)

wb = load_workbook(filename)


try:
    if 'fut7' or 'fut9' or 'fut11' in wb.sheetnames:
        with pd.ExcelWriter("jogos_excel.xlsx", engine="openpyxl", mode = "a", if_sheet_exists = 'replace') as writer:
            interface.prepare_to_excel().to_excel(writer, sheet_name='fut7')
            interface.prepare_to_excel().to_excel(writer, sheet_name='fut9')
            interface.prepare_to_excel().to_excel(writer, sheet_name='fu11')
        
    else:
        with pd.ExcelWriter(filename,engine='openpyxl', mode='a') as writer: 
            interface.prepare_to_excel().to_excel(writer, sheet_name='fut7')
            interface.prepare_to_excel().to_excel(writer, sheet_name='fut9')
            interface.prepare_to_excel().to_excel(writer, sheet_name='fu11')
except Exception as e:
    print(e)



