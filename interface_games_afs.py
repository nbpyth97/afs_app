import tkinter as tk
from tkinter import *
from tkcalendar import Calendar
import pandas as pd
import webscrapping
from tkinter import ttk
from tkinter import scrolledtext
from tkinter.tix import *
import time
import openpyxl
import numpy as np
from openpyxl import load_workbook
import os

#----------Windows features#----------
root = tk.Tk()
root.title("Relatório de Despesas - AFS")
root.geometry("650x650")

newwin = tk.Toplevel(root)
newwin.title('Relatório de Despesas - Download PDF')
newwin.geometry('400x450')
newwin.grab_set()
    
#------------SECONDARY WINDOW FUNCTIONS------------
def grad_date():
    date.config(text = "Selected Date is: " + cal.get_date())

def url_link_sec(url_sec):
    import webbrowser
    webbrowser.open_new(url_sec)
    
def import_chrome_searcher_decide():
    data = cal.get_date()
    if webscrapping.chrome_searcher_nomecao_decide(data.split('-')[0]) == 'Done':
        newwin.grab_release()
        newwin.destroy()
    else:
        date.config(text = "Data Incorreta")
    
def import_chrome_searcher_ultima_nomeacao():
    if webscrapping.chrome_searcher_last_nomeacao() == 'Done':
        newwin.grab_release()
        newwin.destroy()
    else:
        date.config(text = "Data Incorreta")
            
def import_webscrapping():
    try:
        if r2:
            print('O dia facultativo foi selecionado')
            import_chrome_searcher_decide()
    except:
        if r1:
            print('o da ultima nomeação foi selecionado')
            import_chrome_searcher_ultima_nomeacao()
    
def changeState_sec_r1():
    global r1
    if selected_option_r1:
        r1 = 'selected_r1'
        print("selected r1")
        try:
            cal['state']='disabled'
        except Exception as e:
            print(e)
            
def changeState_sec_r2():
    global r2
    if selected_option_r2:
        r2 = 'selected_r2'
        print("selected r2")
        try:
            cal['state']='normal'
        except Exception as e:
            print(e)


#------------SECONDARY WINDOW CODE------------:
selected_option_r1 = IntVar()
selected_option_r2 = IntVar()

cal = Calendar(newwin, selectmode = 'day',date_pattern="dd-mm-yyyy",state="disabled")
cal.place(x=80,y=180)

date = Label(newwin, text = "")
date.place(x=135,y=400)

Label(newwin,text='Última Nomeação:').place(x=10,y=10)
R1_secondary = Radiobutton(newwin, text="Sim", variable=selected_option_r1, value=1,command=changeState_sec_r1)
R1_secondary.place(x=120,y=10)
R2_secondary = Radiobutton(newwin, text="Não", variable=selected_option_r2, value=2,command=changeState_sec_r2)
R2_secondary.place(x=120,y=35)

Label(newwin,text="NOTA IMPORTANTE:\n\n Selecionar a data quando foi colocado o PDF no site da AFS").place(x=20,y=100)

link = Label(newwin, text='AFSetúbal',fg='blue',cursor='hand2')
link.place(x=340,y=0)
link.bind("<Button-1>", lambda e: url_link_sec("https://afsetubal.fpf.pt/Associa%C3%A7%C3%A3o/Documenta%C3%A7%C3%A3o/Nomea%C3%A7%C3%B5es-de-%C3%81rbitros"))

newwin.resizable(False,False)

Button(newwin, text = "Prosseguir",command=import_webscrapping).place(x=320,y=410)


#------------MAIN WINDOW FUNCTIONS------------
lista_inf_jogos_principal=[]
lista_inf_todos_os_jogos=[]

my_entries_1_principal = []
my_entries_2_principal = []
my_entries_3_principal = []
my_entries_4_principal = []
my_entries_5_principal = []
my_entries_6_principal = []
my_entries_7_principal = []
my_entries_8_principal = []

my_entries_1_todos_jogos = []
my_entries_2_todos_jogos = []
my_entries_3_todos_jogos = []
my_entries_4_todos_jogos = []
my_entries_5_todos_jogos = []
my_entries_6_todos_jogos = []
my_entries_7_todos_jogos = []
my_entries_8_todos_jogos = []


my_entries_kms_a = []
my_entries_kms_aa1 = []
my_entries_kms_aa2 = []
kms_label_a = []
kms_label_aa1 = []
kms_label_aa2 = []
arbitro2 = []
team_A_team_B = []
de = []
para = []
nome_do_arbitro = []
aa1 = []
aa2 = []

def export_info_to_excel():
    time.sleep(0.8)

    dataframe = prepare_to_excel()
    filename_main = "{}\jogos_excel.xlsx".format(os.getcwd())
    wb_main = load_workbook(filename_main)
    filename_secondary = "{}\jogos_excel_testing.xlsx".format(os.getcwd())

    try:
        if 'fut' or 'fut_2' in wb_main.sheetnames:
            try:
                if dataframe['2ºÁrbitro']:
                    with pd.ExcelWriter("jogos_excel.xlsx", engine="openpyxl", mode = "a", if_sheet_exists = 'replace') as writer:
                        dataframe.to_excel(writer, sheet_name='fut_2')
                    with pd.ExcelWriter("jogos_excel_testing.xlsx", engine="openpyxl", mode = "a", if_sheet_exists = 'replace') as writer:
                        dataframe.to_excel(writer, sheet_name='fut_2')
            except:
                    with pd.ExcelWriter("jogos_excel.xlsx", engine="openpyxl", mode = "a", if_sheet_exists = 'replace') as writer:
                        dataframe.to_excel(writer, sheet_name='fut')
                    with pd.ExcelWriter("jogos_excel_testing.xlsx", engine="openpyxl", mode = "a", if_sheet_exists = 'replace') as writer:
                        dataframe.to_excel(writer, sheet_name='fut')
        else:
            try:
                if dataframe['2ºÁrbitro']:              
                    with pd.ExcelWriter(filename_main,engine='openpyxl', mode='a') as writer: 
                        dataframe.to_excel(writer, sheet_name='fut_2')
                    with pd.ExcelWriter(filename_secondary,engine='openpyxl', mode='a') as writer: 
                        dataframe.to_excel(writer, sheet_name='fut_2')
            except:
                    with pd.ExcelWriter(filename_main,engine='openpyxl', mode='a') as writer: 
                        dataframe.to_excel(writer, sheet_name='fut')
                    with pd.ExcelWriter(filename_secondary,engine='openpyxl', mode='a') as writer: 
                        dataframe.to_excel(writer, sheet_name='fut')         

    except Exception as e:
        print(e)

def prepare_to_excel():

    time.sleep(0.2)

    if lista_inf_jogos_principal:
        new_mygames = arbitro_principal()

        new_mygames["de A"] = ""
        new_mygames["para A"] = ""
        new_mygames["de AA1"] = ""
        new_mygames["para AA1"] = ""
        new_mygames["de AA2"] = ""
        new_mygames["para AA2"] = ""
        new_mygames['km A'] = ""
        new_mygames['km AA1'] = ""
        new_mygames['km AA2'] = ""
        new_mygames['Itinerário'] = ""

        for num in range(len(new_mygames)):
            viagem = new_mygames.columns.get_loc('de A')
            for i in range(len(lista_inf_jogos_principal)):
                if num+i>len(lista_inf_jogos_principal):
                    break
                else:
                    if i % (len(new_mygames)) == 0:
                        new_mygames.iloc[num,viagem] = lista_inf_jogos_principal[num+i]
                        viagem+=1

        print(new_mygames)
        return new_mygames

    if lista_inf_todos_os_jogos:
        new_mygames_todos = Todos_os_Jogos()

        new_mygames_todos["de A"] = ""
        new_mygames_todos["para A"] = ""
        new_mygames_todos["de AA1"] = ""
        new_mygames_todos["para AA1"] = ""
        new_mygames_todos["de AA2"] = ""
        new_mygames_todos["para AA2"] = ""
        new_mygames_todos['km A'] = ""
        new_mygames_todos['km AA1'] = ""
        new_mygames_todos['km AA2'] = ""
        new_mygames_todos['Itinerário'] = ""

        for num in range(len(new_mygames_todos)):
            viagem = new_mygames_todos.columns.get_loc('de A')
            for i in range(len(lista_inf_todos_os_jogos)):
                if num+i>len(lista_inf_todos_os_jogos):
                    break
                else:
                    if i % (len(new_mygames_todos)) == 0:
                        new_mygames_todos.iloc[num,viagem] = lista_inf_todos_os_jogos[num+i]
                        viagem+=1

        print(new_mygames_todos)
        return new_mygames_todos


def lista_of_values():
    global my_entries_1_principal,my_entries_2_principal,my_entries_3_principal,lista_inf_jogos_principal,lista_inf_todos_os_jogos

    if lista_inf_jogos_principal: 
        if selected_option.get() == 1:
            lista_inf_jogos_principal = []
        print("cleared")
    if lista_inf_todos_os_jogos:
        if selected_option.get() == 2:
            lista_inf_todos_os_jogos = []
        print("cleared")

    else:
        if selected_option.get() == 1:
            for entry_1 in my_entries_1_principal:
                lista_inf_jogos_principal.append(entry_1.get())
            for entry_2 in my_entries_2_principal:
                lista_inf_jogos_principal.append(entry_2.get())
            for entry_4 in my_entries_4_principal:
                lista_inf_jogos_principal.append(entry_4.get())
            for entry_5 in my_entries_5_principal:
                lista_inf_jogos_principal.append(entry_5.get())
            for entry_6 in my_entries_6_principal:
                lista_inf_jogos_principal.append(entry_6.get())
            for entry_7 in my_entries_7_principal:
                lista_inf_jogos_principal.append(entry_7.get())
            for entry_8 in my_entries_8_principal:
                lista_inf_jogos_principal.append(entry_8.get())
            for entry_km in my_entries_kms_a:
                lista_inf_jogos_principal.append(entry_km.get())
            for entry_km in my_entries_kms_aa1:
                lista_inf_jogos_principal.append(entry_km.get())
            for entry_km in my_entries_kms_aa2:
                lista_inf_jogos_principal.append(entry_km.get())
            for entry_3 in my_entries_3_principal:
                entry_text = entry_3.get('1.0', tk.END).split('\n')[0].split('\t')[0]
                lista_inf_jogos_principal.append(entry_text)
            
            print('lista_inf_jogos_principal:',lista_inf_jogos_principal)

        if selected_option.get() == 2:
            for entry_1 in my_entries_1_todos_jogos:
                lista_inf_todos_os_jogos.append(entry_1.get())
            for entry_2 in my_entries_2_todos_jogos:
                lista_inf_todos_os_jogos.append(entry_2.get())
            for entry_4 in my_entries_4_todos_jogos:
                lista_inf_todos_os_jogos.append(entry_4.get())
            for entry_5 in my_entries_5_todos_jogos:
                lista_inf_todos_os_jogos.append(entry_5.get())
            for entry_6 in my_entries_6_todos_jogos:
                lista_inf_todos_os_jogos.append(entry_6.get())
            for entry_7 in my_entries_7_todos_jogos:
                lista_inf_todos_os_jogos.append(entry_7.get())
            for entry_8 in my_entries_8_todos_jogos:
                lista_inf_todos_os_jogos.append(entry_8.get())
            for entry_km in my_entries_kms_a:
                lista_inf_todos_os_jogos.append(entry_km.get())
            for entry_km in my_entries_kms_aa1:
                print(entry_km.get())
                lista_inf_todos_os_jogos.append(entry_km.get())
            for entry_km in my_entries_kms_aa2:
                lista_inf_todos_os_jogos.append(entry_km.get())
            for entry_3 in my_entries_3_todos_jogos:
                entry_text = entry_3.get('1.0',tk.END).split('\n')[0].split('\t')[0]
                lista_inf_todos_os_jogos.append(entry_text)

            print('lista_inf_todos_os_jogos:',lista_inf_todos_os_jogos)

def add_entry_principal():
    global my_entries_1_principal,my_entries_2_principal,my_entries_3_principal,my_entries_kms_a,kms_principal
    new_mygames = arbitro_principal()
    try:
        if my_entries_1_todos_jogos:
            for i in my_entries_1_todos_jogos:
                i.destroy()
            for i in my_entries_2_todos_jogos:
                i.destroy()
            for i in my_entries_3_todos_jogos:
                i.destroy()
            for i in my_entries_kms_a:
                i.destroy()
            for i in my_entries_kms_aa1:
                i.destroy()
            for i in my_entries_kms_aa2:
                i.destroy()
            for i in kms_label_a:
                i.destroy()    
            for i in kms_label_aa1:
                i.destroy()  
            for i in kms_label_aa2:
                i.destroy()  
            for i in kms_label_a:
                i.destroy()
            for i in team_A_team_B:
                i.destroy() 
            for i in my_entries_4_todos_jogos:
                i.destroy()
            for i in my_entries_5_todos_jogos:
                i.destroy()
            for i in my_entries_6_todos_jogos:
                i.destroy()
            for i in my_entries_7_todos_jogos:
                i.destroy()
            for i in my_entries_8_todos_jogos:
                i.destroy()
            for i in de:
                i.destroy()
            for i in para:
                i.destroy()
            for i in nome_do_arbitro:
                i.destroy()
            for i in aa1:
                i.destroy()
            for i in aa2:
                i.destroy()
        
            count = 1 
            MAX_NUM = len(new_mygames)
            x_val_1 = 140
            y_val_entry_1 = 160
            y_val_entry_2 = 190
            y_val_entry_3 = 220
            y_val_entry_scroll = 150

            for count in range(MAX_NUM):
                if count == 3:
                    x_val_1 = 140
                    y_val_entry_1 = 160
                    y_val_entry_2 = 190
                    y_val_entry_3 = 220
                    y_val_entry_scroll = 150
                if count > 2:
                    root.geometry("1200x1400")
                    root.resizable(True,True)
                    print(new_mygames.iloc[count]['Equipa A'],'-',new_mygames.iloc[count]['Equipa B'])

                    my_entries_1_principal.append(Entry(root))
                    my_entries_4_principal.append(Entry(root))
                    my_entries_7_principal.append(Entry(root))

                    my_entries_2_principal.append(Entry(root))
                    my_entries_5_principal.append(Entry(root))
                    my_entries_8_principal.append(Entry(root))
                    
                    #DE:
                    my_entries_1_principal[-1].place(x=x_val_1+800,y=y_val_entry_1)
                    my_entries_4_principal[-1].place(x=x_val_1+800,y=y_val_entry_2)
                    my_entries_7_principal[-1].place(x=x_val_1+800,y=y_val_entry_3)

                    #PARA:
                    my_entries_2_principal[-1].place(x=x_val_1+150+800,y=y_val_entry_1)
                    my_entries_5_principal[-1].place(x=x_val_1+150+800,y=y_val_entry_2)
                    my_entries_8_principal[-1].place(x=x_val_1+150+800,y=y_val_entry_3)

                    #ITENERARIO:
                    my_entries_3_principal.append(scrolledtext.ScrolledText(root,wrap=tk.WORD,width=19,height=6))
                    my_entries_3_principal[-1].place(x=x_val_1+300+800,y=y_val_entry_scroll)

                    my_entries_kms_a.append(Entry(root,width=6))
                    my_entries_kms_aa1.append(Entry(root,width=6))
                    my_entries_kms_aa2.append(Entry(root,width=6))
                    my_entries_kms_a[-1].place(x=x_val_1+810,y=y_val_entry_3+30)
                    my_entries_kms_aa1[-1].place(x=x_val_1+810+100,y=y_val_entry_3+30)
                    my_entries_kms_aa2[-1].place(x=x_val_1+810+200,y=y_val_entry_3+30)

                    kms_label_a.append(Label(text='KMS A:'))
                    kms_label_aa1.append(Label(text='KMS AA1:'))
                    kms_label_aa2.append(Label(text='KMS AA2:'))
                    kms_label_a[-1].place(x=x_val_1+700,y=y_val_entry_3+30)
                    kms_label_aa1[-1].place(x=x_val_1+745+100,y=y_val_entry_3+30)
                    kms_label_aa2[-1].place(x=x_val_1+745+200,y=y_val_entry_3+30)

                    de.append(Label(text='DE:'))
                    de[-1].place(x=x_val_1+800,y=y_val_entry_1-25)
                    para.append(Label(text='PARA:'))
                    para[-1].place(x=x_val_1+150+800,y=y_val_entry_1-25)

                    team_A_team_B.append(Label(text='{}-{}:'.format(new_mygames.iloc[count]['Equipa A'],new_mygames.iloc[count]['Equipa B'])))
                    team_A_team_B[-1].place(x=x_val_1+20+800,y=y_val_entry_1-50)
                    nome_do_arbitro.append(Label(text='{}:'.format(new_mygames.iloc[count]['Árbitro'])))
                    nome_do_arbitro[-1].place(x=5+800,y=y_val_entry_1)
                    aa1.append(Label(text='{}:'.format(new_mygames.iloc[count]['AA1'])))
                    aa1[-1].place(x=5+800,y=y_val_entry_2)
                    aa2.append(Label(text='{}:'.format(new_mygames.iloc[count]['AA2'])))
                    aa2[-1].place(x=5+800,y=y_val_entry_3)

                    if isinstance(new_mygames.iloc[count]['AA1'],float):
                        try:
                            arbitro2.append(Label(text='{}:'.format(new_mygames.iloc[count]['2ºÁrbitro'])))
                            arbitro2[-1].place(x=5,y=y_val_entry_2)
                        except:
                            pass

                    y_val_entry_1 = y_val_entry_1 + 170
                    y_val_entry_2 = y_val_entry_2 + 170
                    y_val_entry_3 = y_val_entry_3 + 170
                    y_val_entry_scroll = y_val_entry_scroll + 170
                    count += 1
                else:    
                    print(new_mygames.iloc[count]['Equipa A'],'-',new_mygames.iloc[count]['Equipa B'])
                    my_entries_1_principal.append(Entry(root))
                    my_entries_4_principal.append(Entry(root))
                    my_entries_7_principal.append(Entry(root))

                    my_entries_2_principal.append(Entry(root))
                    my_entries_5_principal.append(Entry(root))
                    my_entries_8_principal.append(Entry(root))
                    
                    #DE:
                    my_entries_1_principal[-1].place(x=x_val_1,y=y_val_entry_1)
                    my_entries_4_principal[-1].place(x=x_val_1,y=y_val_entry_2)
                    my_entries_7_principal[-1].place(x=x_val_1,y=y_val_entry_3)

                    #PARA:
                    my_entries_2_principal[-1].place(x=x_val_1+150,y=y_val_entry_1)
                    my_entries_5_principal[-1].place(x=x_val_1+150,y=y_val_entry_2)
                    my_entries_8_principal[-1].place(x=x_val_1+150,y=y_val_entry_3)

                    #ITENERARIO:
                    my_entries_3_principal.append(scrolledtext.ScrolledText(root,wrap=tk.WORD,width=19,height=6))
                    my_entries_3_principal[-1].place(x=x_val_1+300,y=y_val_entry_scroll)

                    my_entries_kms_a.append(Entry(root,width=6))
                    my_entries_kms_aa1.append(Entry(root,width=6))
                    my_entries_kms_aa2.append(Entry(root,width=6))
                    my_entries_kms_a[-1].place(x=x_val_1,y=y_val_entry_3+30)
                    my_entries_kms_aa1[-1].place(x=x_val_1+110,y=y_val_entry_3+30)
                    my_entries_kms_aa2[-1].place(x=x_val_1+210,y=y_val_entry_3+30)

                    kms_label_a.append(Label(text='KMS A:'))
                    kms_label_aa1.append(Label(text='KMS AA1:'))
                    kms_label_aa2.append(Label(text='KMS AA2:'))
                    kms_label_a[-1].place(x=x_val_1-50,y=y_val_entry_3+30)
                    kms_label_aa1[-1].place(x=x_val_1+45,y=y_val_entry_3+30)
                    kms_label_aa2[-1].place(x=x_val_1+145,y=y_val_entry_3+30)

                    de.append(Label(text='DE:'))
                    de[-1].place(x=x_val_1,y=y_val_entry_1-25)
                    para.append(Label(text='PARA:'))
                    para[-1].place(x=x_val_1+150,y=y_val_entry_1-25)

                    team_A_team_B.append(Label(text='{}-{}:'.format(new_mygames.iloc[count]['Equipa A'],new_mygames.iloc[count]['Equipa B'])))
                    team_A_team_B[-1].place(x=x_val_1+20,y=y_val_entry_1-50)
                    nome_do_arbitro.append(Label(text='{}:'.format(new_mygames.iloc[count]['Árbitro'])))
                    nome_do_arbitro[-1].place(x=5,y=y_val_entry_1)
                    aa1.append(Label(text='{}:'.format(new_mygames.iloc[count]['AA1'])))
                    aa1[-1].place(x=5,y=y_val_entry_2)
                    aa2.append(Label(text='{}:'.format(new_mygames.iloc[count]['AA2'])))
                    aa2[-1].place(x=5,y=y_val_entry_3)

                    if isinstance(new_mygames.iloc[count]['AA1'],float):
                        try:
                            arbitro2.append(Label(text='{}:'.format(new_mygames.iloc[count]['2ºÁrbitro'])))
                            arbitro2[-1].place(x=5,y=y_val_entry_2)
                        except:
                            pass


                    y_val_entry_1 = y_val_entry_1 + 170
                    y_val_entry_2 = y_val_entry_2 + 170
                    y_val_entry_3 = y_val_entry_3 + 170
                    y_val_entry_scroll = y_val_entry_scroll + 170
                    count += 1

        else:
            count = 1 
            MAX_NUM = len(new_mygames)
            x_val_1 = 140
            y_val_entry_1 = 160
            y_val_entry_2 = 190
            y_val_entry_3 = 220
            y_val_entry_scroll = 150
            for count in range(MAX_NUM):
                if count == 3:
                    x_val_1 = 140
                    y_val_entry_1 = 160
                    y_val_entry_2 = 190
                    y_val_entry_3 = 220
                    y_val_entry_scroll = 150
                if count > 2:
                    root.geometry("1200x1400")
                    root.resizable(True,True)
                    print(new_mygames.iloc[count]['Equipa A'],'-',new_mygames.iloc[count]['Equipa B'])

                    my_entries_1_principal.append(Entry(root))
                    my_entries_4_principal.append(Entry(root))
                    my_entries_7_principal.append(Entry(root))

                    my_entries_2_principal.append(Entry(root))
                    my_entries_5_principal.append(Entry(root))
                    my_entries_8_principal.append(Entry(root))
                    
                    #DE:
                    my_entries_1_principal[-1].place(x=x_val_1+800,y=y_val_entry_1)
                    my_entries_4_principal[-1].place(x=x_val_1+800,y=y_val_entry_2)
                    my_entries_7_principal[-1].place(x=x_val_1+800,y=y_val_entry_3)

                    #PARA:
                    my_entries_2_principal[-1].place(x=x_val_1+150+800,y=y_val_entry_1)
                    my_entries_5_principal[-1].place(x=x_val_1+150+800,y=y_val_entry_2)
                    my_entries_8_principal[-1].place(x=x_val_1+150+800,y=y_val_entry_3)

                    #ITENERARIO:
                    my_entries_3_principal.append(scrolledtext.ScrolledText(root,wrap=tk.WORD,width=19,height=6))
                    my_entries_3_principal[-1].place(x=x_val_1+300+800,y=y_val_entry_scroll)

                    my_entries_kms_a.append(Entry(root,width=6))
                    my_entries_kms_aa1.append(Entry(root,width=6))
                    my_entries_kms_aa2.append(Entry(root,width=6))
                    my_entries_kms_a[-1].place(x=x_val_1+810,y=y_val_entry_3+30)
                    my_entries_kms_aa1[-1].place(x=x_val_1+810+100,y=y_val_entry_3+30)
                    my_entries_kms_aa2[-1].place(x=x_val_1+810+200,y=y_val_entry_3+30)

                    kms_label_a.append(Label(text='KMS A:'))
                    kms_label_aa1.append(Label(text='KMS AA1:'))
                    kms_label_aa2.append(Label(text='KMS AA2:'))
                    kms_label_a[-1].place(x=x_val_1+700,y=y_val_entry_3+30)
                    kms_label_aa1[-1].place(x=x_val_1+745+100,y=y_val_entry_3+30)
                    kms_label_aa2[-1].place(x=x_val_1+745+200,y=y_val_entry_3+30)

                    de.append(Label(text='DE:'))
                    de[-1].place(x=x_val_1+800,y=y_val_entry_1-25)
                    para.append(Label(text='PARA:'))
                    para[-1].place(x=x_val_1+150+800,y=y_val_entry_1-25)

                    team_A_team_B.append(Label(text='{}-{}:'.format(new_mygames.iloc[count]['Equipa A'],new_mygames.iloc[count]['Equipa B'])))
                    team_A_team_B[-1].place(x=x_val_1+20+800,y=y_val_entry_1-50)
                    nome_do_arbitro.append(Label(text='{}:'.format(new_mygames.iloc[count]['Árbitro'])))
                    nome_do_arbitro[-1].place(x=5+800,y=y_val_entry_1)
                    aa1.append(Label(text='{}:'.format(new_mygames.iloc[count]['AA1'])))
                    aa1[-1].place(x=5+800,y=y_val_entry_2)
                    aa2.append(Label(text='{}:'.format(new_mygames.iloc[count]['AA2'])))
                    aa2[-1].place(x=5+800,y=y_val_entry_3)

                    if isinstance(new_mygames.iloc[count]['AA1'],float):
                        try:
                            arbitro2.append(Label(text='{}:'.format(new_mygames.iloc[count]['2ºÁrbitro'])))
                            arbitro2[-1].place(x=5,y=y_val_entry_2)
                        except:
                            pass

                    y_val_entry_1 = y_val_entry_1 + 170
                    y_val_entry_2 = y_val_entry_2 + 170
                    y_val_entry_3 = y_val_entry_3 + 170
                    y_val_entry_scroll = y_val_entry_scroll + 170
                    count += 1
                else:
                    print(new_mygames.iloc[count]['Equipa A'],'-',new_mygames.iloc[count]['Equipa B'])
                    my_entries_1_principal.append(Entry(root))
                    my_entries_4_principal.append(Entry(root))
                    my_entries_7_principal.append(Entry(root))

                    my_entries_2_principal.append(Entry(root))
                    my_entries_5_principal.append(Entry(root))
                    my_entries_8_principal.append(Entry(root))
                    
                    #DE:
                    my_entries_1_principal[-1].place(x=x_val_1,y=y_val_entry_1)
                    my_entries_4_principal[-1].place(x=x_val_1,y=y_val_entry_2)
                    my_entries_7_principal[-1].place(x=x_val_1,y=y_val_entry_3)

                    #PARA:
                    my_entries_2_principal[-1].place(x=x_val_1+150,y=y_val_entry_1)
                    my_entries_5_principal[-1].place(x=x_val_1+150,y=y_val_entry_2)
                    my_entries_8_principal[-1].place(x=x_val_1+150,y=y_val_entry_3)

                    #ITENERARIO:
                    my_entries_3_principal.append(scrolledtext.ScrolledText(root,wrap=tk.WORD,width=19,height=6))
                    my_entries_3_principal[-1].place(x=x_val_1+300,y=y_val_entry_scroll)

                    my_entries_kms_a.append(Entry(root,width=6))
                    my_entries_kms_aa1.append(Entry(root,width=6))
                    my_entries_kms_aa2.append(Entry(root,width=6))
                    my_entries_kms_a[-1].place(x=x_val_1,y=y_val_entry_3+30)
                    my_entries_kms_aa1[-1].place(x=x_val_1+110,y=y_val_entry_3+30)
                    my_entries_kms_aa2[-1].place(x=x_val_1+210,y=y_val_entry_3+30)

                    kms_label_a.append(Label(text='KMS A:'))
                    kms_label_aa1.append(Label(text='KMS AA1:'))
                    kms_label_aa2.append(Label(text='KMS AA2:'))
                    kms_label_a[-1].place(x=x_val_1-50,y=y_val_entry_3+30)
                    kms_label_aa1[-1].place(x=x_val_1+45,y=y_val_entry_3+30)
                    kms_label_aa2[-1].place(x=x_val_1+145,y=y_val_entry_3+30)

                    de.append(Label(text='DE:'))
                    de[-1].place(x=x_val_1,y=y_val_entry_1-25)
                    para.append(Label(text='PARA:'))
                    para[-1].place(x=x_val_1+150,y=y_val_entry_1-25)

                    team_A_team_B.append(Label(text='{}-{}:'.format(new_mygames.iloc[count]['Equipa A'],new_mygames.iloc[count]['Equipa B'])))
                    team_A_team_B[-1].place(x=x_val_1+20,y=y_val_entry_1-50)
                    nome_do_arbitro.append(Label(text='{}:'.format(new_mygames.iloc[count]['Árbitro'])))
                    nome_do_arbitro[-1].place(x=5,y=y_val_entry_1)
                    aa1.append(Label(text='{}:'.format(new_mygames.iloc[count]['AA1'])))
                    aa1[-1].place(x=5,y=y_val_entry_2)
                    aa2.append(Label(text='{}:'.format(new_mygames.iloc[count]['AA2'])))
                    aa2[-1].place(x=5,y=y_val_entry_3)

                    if isinstance(new_mygames.iloc[count]['AA1'],float):
                        try:
                            arbitro2.append(Label(text='{}:'.format(new_mygames.iloc[count]['2ºÁrbitro'])))
                            arbitro2[-1].place(x=5,y=y_val_entry_2)
                        except:
                            pass

                    y_val_entry_1 = y_val_entry_1 + 170
                    y_val_entry_2 = y_val_entry_2 + 170
                    y_val_entry_3 = y_val_entry_3 + 170
                    y_val_entry_scroll = y_val_entry_scroll + 170
                    count += 1
    except Exception as e:
        print(e)

        return my_entries_1_principal,my_entries_2_principal,my_entries_3_principal

def add_entry_todos():
    global my_entries_1_todos_jogos,my_entries_2_todos_jogos,my_entries_3_todos_jogos,my_entries_kms_a,kms_todos_jogos
    new_mygames_todos = Todos_os_Jogos()
    try:
        if my_entries_1_principal:
            for i in my_entries_1_principal:
                i.destroy()
            for i in my_entries_2_principal:
                i.destroy()
            for i in my_entries_3_principal:
                i.destroy()
            for i in my_entries_kms_a:
                i.destroy()
            for i in my_entries_kms_aa1:
                i.destroy()
            for i in my_entries_kms_aa2:
                i.destroy()
            for i in kms_label_a:
                i.destroy()    
            for i in kms_label_aa1:
                i.destroy()  
            for i in kms_label_aa2:
                i.destroy()  
            for i in team_A_team_B:
                i.destroy()  
            for i in my_entries_4_principal:
                i.destroy()
            for i in my_entries_5_principal:
                i.destroy()
            for i in my_entries_6_principal:
                i.destroy()
            for i in my_entries_7_principal:
                i.destroy()
            for i in my_entries_8_principal:
                i.destroy()
            for i in de:
                i.destroy()
            for i in para:
                i.destroy()
            for i in nome_do_arbitro:
                i.destroy()
            for i in aa1:
                i.destroy()
            for i in aa2:
                i.destroy()

            count = 1 
            MAX_NUM = len(new_mygames_todos)
            x_val_1 = 140
            y_val_entry_1 = 160
            y_val_entry_2 = 190
            y_val_entry_3 = 220
            y_val_entry_scroll = 150
            for count in range(MAX_NUM):
                if count == 3:
                    x_val_1 = 140
                    y_val_entry_1 = 160
                    y_val_entry_2 = 190
                    y_val_entry_3 = 220
                    y_val_entry_scroll = 150
                if count > 2:
                    root.geometry("1200x1400")
                    root.resizable(True,True)
                    print(new_mygames.iloc[count]['Equipa A'],'-',new_mygames.iloc[count]['Equipa B'])
                    my_entries_1_todos_jogos.append(Entry(root))
                    my_entries_4_todos_jogos.append(Entry(root))
                    my_entries_7_todos_jogos.append(Entry(root))

                    my_entries_2_todos_jogos.append(Entry(root))
                    my_entries_5_todos_jogos.append(Entry(root))
                    my_entries_8_todos_jogos.append(Entry(root))
                    
                    #DE:
                    my_entries_1_todos_jogos[-1].place(x=x_val_1+800,y=y_val_entry_1)
                    my_entries_4_todos_jogos[-1].place(x=x_val_1+800,y=y_val_entry_2)
                    my_entries_7_todos_jogos[-1].place(x=x_val_1+800,y=y_val_entry_3)

                    #PARA:
                    my_entries_2_todos_jogos[-1].place(x=x_val_1+800+150,y=y_val_entry_1)
                    my_entries_5_todos_jogos[-1].place(x=x_val_1+800+150,y=y_val_entry_2)
                    my_entries_8_todos_jogos[-1].place(x=x_val_1+800+150,y=y_val_entry_3)

                    #ITENERARIO:
                    my_entries_3_todos_jogos.append(scrolledtext.ScrolledText(root,wrap=tk.WORD,width=19,height=6))
                    my_entries_3_todos_jogos[-1].place(x=x_val_1+800+300,y=y_val_entry_scroll)

                    my_entries_kms_a.append(Entry(root,width=6))
                    my_entries_kms_aa1.append(Entry(root,width=6))
                    my_entries_kms_aa2.append(Entry(root,width=6))
                    my_entries_kms_a[-1].place(x=x_val_1+810,y=y_val_entry_3+30)
                    my_entries_kms_aa1[-1].place(x=x_val_1+810+100,y=y_val_entry_3+30)
                    my_entries_kms_aa2[-1].place(x=x_val_1+810+200,y=y_val_entry_3+30)

                    kms_label_a.append(Label(text='KMS A:'))
                    kms_label_aa1.append(Label(text='KMS AA1:'))
                    kms_label_aa2.append(Label(text='KMS AA2:'))
                    kms_label_a[-1].place(x=x_val_1+700,y=y_val_entry_3+30)
                    kms_label_aa1[-1].place(x=x_val_1+745+100,y=y_val_entry_3+30)
                    kms_label_aa2[-1].place(x=x_val_1+745+200,y=y_val_entry_3+30)

                    de.append(Label(text='DE:'))
                    de[-1].place(x=x_val_1+800,y=y_val_entry_1-25)
                    para.append(Label(text='PARA:'))
                    para[-1].place(x=x_val_1+800+150,y=y_val_entry_1-25)

                    team_A_team_B.append(Label(text='{}-{}:'.format(new_mygames.iloc[count]['Equipa A'],new_mygames.iloc[count]['Equipa B'])))
                    team_A_team_B[-1].place(x=x_val_1+800+20,y=y_val_entry_1-50)
                    nome_do_arbitro.append(Label(text='{}:'.format(new_mygames.iloc[count]['Árbitro'])))
                    nome_do_arbitro[-1].place(x=5+800,y=y_val_entry_1)
                    aa1.append(Label(text='{}:'.format(new_mygames.iloc[count]['AA1'])))
                    aa1[-1].place(x=5+800,y=y_val_entry_2)
                    aa2.append(Label(text='{}:'.format(new_mygames.iloc[count]['AA2'])))
                    aa2[-1].place(x=5+800,y=y_val_entry_3)

                    if isinstance(new_mygames.iloc[count]['AA1'],float):
                        arbitro2.append(Label(text='{}:'.format(new_mygames.iloc[count]['2ºÁrbitro'])))
                        arbitro2[-1].place(x=5,y=y_val_entry_2)

                    y_val_entry_1 = y_val_entry_1 + 170
                    y_val_entry_2 = y_val_entry_2 + 170
                    y_val_entry_3 = y_val_entry_3 + 170
                    y_val_entry_scroll = y_val_entry_scroll + 170
                    count += 1
                else:
                    print(new_mygames.iloc[count-1]['Equipa A'],'-',new_mygames.iloc[count]['Equipa B'])
                    my_entries_1_todos_jogos.append(Entry(root))
                    my_entries_4_todos_jogos.append(Entry(root))
                    my_entries_7_todos_jogos.append(Entry(root))

                    my_entries_2_todos_jogos.append(Entry(root))
                    my_entries_5_todos_jogos.append(Entry(root))
                    my_entries_8_todos_jogos.append(Entry(root))
                    
                    #DE:
                    my_entries_1_todos_jogos[-1].place(x=x_val_1,y=y_val_entry_1)
                    my_entries_4_todos_jogos[-1].place(x=x_val_1,y=y_val_entry_2)
                    my_entries_7_todos_jogos[-1].place(x=x_val_1,y=y_val_entry_3)

                    #PARA:
                    my_entries_2_todos_jogos[-1].place(x=x_val_1+150,y=y_val_entry_1)
                    my_entries_5_todos_jogos[-1].place(x=x_val_1+150,y=y_val_entry_2)
                    my_entries_8_todos_jogos[-1].place(x=x_val_1+150,y=y_val_entry_3)

                    #ITENERARIO:
                    my_entries_3_todos_jogos.append(scrolledtext.ScrolledText(root,wrap=tk.WORD,width=19,height=6))
                    my_entries_3_todos_jogos[-1].place(x=x_val_1+300,y=y_val_entry_scroll)

                    my_entries_kms_a.append(Entry(root,width=6))
                    my_entries_kms_aa1.append(Entry(root,width=6))
                    my_entries_kms_aa2.append(Entry(root,width=6))
                    my_entries_kms_a[-1].place(x=x_val_1,y=y_val_entry_3+30)
                    my_entries_kms_aa1[-1].place(x=x_val_1+110,y=y_val_entry_3+30)
                    my_entries_kms_aa2[-1].place(x=x_val_1+210,y=y_val_entry_3+30)

                    kms_label_a.append(Label(text='KMS A:'))
                    kms_label_aa1.append(Label(text='KMS AA1:'))
                    kms_label_aa2.append(Label(text='KMS AA2:'))
                    kms_label_a[-1].place(x=x_val_1-50,y=y_val_entry_3+30)
                    kms_label_aa1[-1].place(x=x_val_1+45,y=y_val_entry_3+30)
                    kms_label_aa2[-1].place(x=x_val_1+145,y=y_val_entry_3+30)

                    de.append(Label(text='DE:'))
                    de[-1].place(x=x_val_1,y=y_val_entry_1-25)
                    para.append(Label(text='PARA:'))
                    para[-1].place(x=x_val_1+150,y=y_val_entry_1-25)

                    team_A_team_B.append(Label(text='{}-{}:'.format(new_mygames.iloc[count]['Equipa A'],new_mygames.iloc[count]['Equipa B'])))
                    team_A_team_B[-1].place(x=x_val_1+20,y=y_val_entry_1-50)
                    nome_do_arbitro.append(Label(text='{}:'.format(new_mygames.iloc[count]['Árbitro'])))
                    nome_do_arbitro[-1].place(x=5,y=y_val_entry_1)
                    aa1.append(Label(text='{}:'.format(new_mygames.iloc[count]['AA1'])))
                    aa1[-1].place(x=5,y=y_val_entry_2)
                    aa2.append(Label(text='{}:'.format(new_mygames.iloc[count]['AA2'])))
                    aa2[-1].place(x=5,y=y_val_entry_3)

                    if isinstance(new_mygames.iloc[count]['AA1'],float):
                        try:
                            arbitro2.append(Label(text='{}:'.format(new_mygames.iloc[count]['2ºÁrbitro'])))
                            arbitro2[-1].place(x=5,y=y_val_entry_2)
                        except:
                            pass

                    y_val_entry_1 = y_val_entry_1 + 170
                    y_val_entry_2 = y_val_entry_2 + 170
                    y_val_entry_3 = y_val_entry_3 + 170
                    y_val_entry_scroll = y_val_entry_scroll + 170
                    count += 1
        else:
            count = 1 
            MAX_NUM = len(new_mygames_todos)
            x_val_1 = 140
            y_val_entry_1 = 160
            y_val_entry_2 = 190
            y_val_entry_3 = 220
            y_val_entry_scroll = 150
            for count in range(MAX_NUM):
                if count == 3:
                    x_val_1 = 140
                    y_val_entry_1 = 160
                    y_val_entry_2 = 190
                    y_val_entry_3 = 220
                    y_val_entry_scroll = 150
                if count > 2:
                    root.geometry("1200x1400")
                    root.resizable(True,True)
                    print(new_mygames.iloc[count]['Equipa A'],'-',new_mygames.iloc[count]['Equipa B'])
                    my_entries_1_todos_jogos.append(Entry(root))
                    my_entries_4_todos_jogos.append(Entry(root))
                    my_entries_7_todos_jogos.append(Entry(root))

                    my_entries_2_todos_jogos.append(Entry(root))
                    my_entries_5_todos_jogos.append(Entry(root))
                    my_entries_8_todos_jogos.append(Entry(root))
                    
                    #DE:
                    my_entries_1_todos_jogos[-1].place(x=x_val_1+800,y=y_val_entry_1)
                    my_entries_4_todos_jogos[-1].place(x=x_val_1+800,y=y_val_entry_2)
                    my_entries_7_todos_jogos[-1].place(x=x_val_1+800,y=y_val_entry_3)

                    #PARA:
                    my_entries_2_todos_jogos[-1].place(x=x_val_1+800+10,y=y_val_entry_1)
                    my_entries_5_todos_jogos[-1].place(x=x_val_1+800+150,y=y_val_entry_2)
                    my_entries_8_todos_jogos[-1].place(x=x_val_1+800+150,y=y_val_entry_3)

                    #ITENERARIO:
                    my_entries_3_todos_jogos.append(scrolledtext.ScrolledText(root,wrap=tk.WORD,width=19,height=6))
                    my_entries_3_todos_jogos[-1].place(x=x_val_1+800+300,y=y_val_entry_scroll)

                    my_entries_kms_a.append(Entry(root,width=6))
                    my_entries_kms_aa1.append(Entry(root,width=6))
                    my_entries_kms_aa2.append(Entry(root,width=6))
                    my_entries_kms_a[-1].place(x=x_val_1+810,y=y_val_entry_3+30)
                    my_entries_kms_aa1[-1].place(x=x_val_1+810+100,y=y_val_entry_3+30)
                    my_entries_kms_aa2[-1].place(x=x_val_1+810+200,y=y_val_entry_3+30)

                    kms_label_a.append(Label(text='KMS A:'))
                    kms_label_aa1.append(Label(text='KMS AA1:'))
                    kms_label_aa2.append(Label(text='KMS AA2:'))
                    kms_label_a[-1].place(x=x_val_1+700,y=y_val_entry_3+30)
                    kms_label_aa1[-1].place(x=x_val_1+745+100,y=y_val_entry_3+30)
                    kms_label_aa2[-1].place(x=x_val_1+745+200,y=y_val_entry_3+30)

                    de.append(Label(text='DE:'))
                    de[-1].place(x=x_val_1+800,y=y_val_entry_1-25)
                    para.append(Label(text='PARA:'))
                    para[-1].place(x=x_val_1+800+150,y=y_val_entry_1-25)

                    team_A_team_B.append(Label(text='{}-{}:'.format(new_mygames.iloc[count]['Equipa A'],new_mygames.iloc[count]['Equipa B'])))
                    team_A_team_B[-1].place(x=x_val_1+800+20,y=y_val_entry_1-50)
                    nome_do_arbitro.append(Label(text='{}:'.format(new_mygames.iloc[count]['Árbitro'])))
                    nome_do_arbitro[-1].place(x=5+800,y=y_val_entry_1)
                    aa1.append(Label(text='{}:'.format(new_mygames.iloc[count]['AA1'])))
                    aa1[-1].place(x=5+800,y=y_val_entry_2)
                    aa2.append(Label(text='{}:'.format(new_mygames.iloc[count]['AA2'])))
                    aa2[-1].place(x=5+800,y=y_val_entry_3)
                    
                    if isinstance(new_mygames.iloc[count]['AA1'],float):
                        try:
                            arbitro2.append(Label(text='{}:'.format(new_mygames.iloc[count]['2ºÁrbitro'])))
                            arbitro2[-1].place(x=5,y=y_val_entry_2)
                        except:
                            pass

                    y_val_entry_1 = y_val_entry_1 + 170
                    y_val_entry_2 = y_val_entry_2 + 170
                    y_val_entry_3 = y_val_entry_3 + 170
                    y_val_entry_scroll = y_val_entry_scroll + 170
                    count += 1
                    
                else:
                    print(new_mygames.iloc[count]['Equipa A'],'-',new_mygames.iloc[count]['Equipa B'])
                    my_entries_1_todos_jogos.append(Entry(root))
                    my_entries_4_todos_jogos.append(Entry(root))
                    my_entries_7_todos_jogos.append(Entry(root))

                    my_entries_2_todos_jogos.append(Entry(root))
                    my_entries_5_todos_jogos.append(Entry(root))
                    my_entries_8_todos_jogos.append(Entry(root))
                    
                    #DE:
                    my_entries_1_todos_jogos[-1].place(x=x_val_1,y=y_val_entry_1)
                    my_entries_4_todos_jogos[-1].place(x=x_val_1,y=y_val_entry_2)
                    my_entries_7_todos_jogos[-1].place(x=x_val_1,y=y_val_entry_3)

                    #PARA:
                    my_entries_2_todos_jogos[-1].place(x=x_val_1+150,y=y_val_entry_1)
                    my_entries_5_todos_jogos[-1].place(x=x_val_1+150,y=y_val_entry_2)
                    my_entries_8_todos_jogos[-1].place(x=x_val_1+150,y=y_val_entry_3)

                    #ITENERARIO:
                    my_entries_3_todos_jogos.append(scrolledtext.ScrolledText(root,wrap=tk.WORD,width=19,height=6))
                    my_entries_3_todos_jogos[-1].place(x=x_val_1+300,y=y_val_entry_scroll)

                    my_entries_kms_a.append(Entry(root,width=6))
                    my_entries_kms_aa1.append(Entry(root,width=6))
                    my_entries_kms_aa2.append(Entry(root,width=6))
                    my_entries_kms_a[-1].place(x=x_val_1,y=y_val_entry_3+30)
                    my_entries_kms_aa1[-1].place(x=x_val_1+100,y=y_val_entry_3+30)
                    my_entries_kms_aa2[-1].place(x=x_val_1+200,y=y_val_entry_3+30)

                    kms_label_a.append(Label(text='KMS A:'))
                    kms_label_aa1.append(Label(text='KMS AA1:'))
                    kms_label_aa2.append(Label(text='KMS AA2:'))
                    kms_label_a[-1].place(x=x_val_1-50,y=y_val_entry_3+30)
                    kms_label_aa1[-1].place(x=x_val_1+45,y=y_val_entry_3+30)
                    kms_label_aa2[-1].place(x=x_val_1+145,y=y_val_entry_3+30)

                    de.append(Label(text='DE:'))
                    de[-1].place(x=x_val_1,y=y_val_entry_1-25)
                    para.append(Label(text='PARA:'))
                    para[-1].place(x=x_val_1+150,y=y_val_entry_1-25)

                    team_A_team_B.append(Label(text='{}-{}:'.format(new_mygames.iloc[count]['Equipa A'],new_mygames.iloc[count]['Equipa B'])))
                    team_A_team_B[-1].place(x=x_val_1+20,y=y_val_entry_1-50)
                    nome_do_arbitro.append(Label(text='{}:'.format(new_mygames.iloc[count]['Árbitro'])))
                    nome_do_arbitro[-1].place(x=5,y=y_val_entry_1)
                    aa1.append(Label(text='{}:'.format(new_mygames.iloc[count]['AA1'])))
                    aa1[-1].place(x=5,y=y_val_entry_2)
                    aa2.append(Label(text='{}:'.format(new_mygames.iloc[count]['AA2'])))
                    aa2[-1].place(x=5,y=y_val_entry_3)
                    
                    if isinstance(new_mygames.iloc[count]['AA1'],float):
                        try:
                            arbitro2.append(Label(text='{}:'.format(new_mygames.iloc[count]['2ºÁrbitro'])))
                            arbitro2[-1].place(x=5,y=y_val_entry_2)
                        except:
                            pass

                    y_val_entry_1 = y_val_entry_1 + 170
                    y_val_entry_2 = y_val_entry_2 + 170
                    y_val_entry_3 = y_val_entry_3 + 170
                    y_val_entry_scroll = y_val_entry_scroll + 170
                    count += 1
    except Exception as e:
        print(e)  
            
    return my_entries_1_todos_jogos,my_entries_2_todos_jogos,my_entries_3_todos_jogos

def km_calculation():
    try:
        print(km_variable)
        result= float(km_variable.get())*float(km_quantity.get())
        result = round(result,3)
    except Exception as e:
        print(e)
    result_km.set(result)   
    
def url_link(url):
    import webbrowser
    webbrowser.open_new(url)
    
def arbitro_principal():
    global new_mygames
    try:
        i=0
        new_mygames = pd.concat([mygames[0],mygames[1],mygames[2]])
        new_mygames = new_mygames.reset_index()
        new_mygames_official = new_mygames.copy()

        #Árbitro Principal:
        for i in range(len(new_mygames_official)):
            if new_mygames_official.iloc[i][2] == '({})'.format(n_arbitro_var.get()):
                pass
            else:
                new_mygames = new_mygames.drop(i)   

        print(new_mygames)

    except Exception as e:
        print(e)
        
    return new_mygames

def Todos_os_Jogos():
    global new_mygames
    new_mygames = pd.concat([mygames[0],mygames[1],mygames[2]])
    new_mygames = new_mygames.reset_index()
    
    print(new_mygames)
    
    return new_mygames

def import_excel_doc():
    global mygames

    import excel_doc_afs as eda
    print('entrei import_excel_doc')
    mygames = eda.jogos_do_arbitro(n_arbitro_var.get())

    return mygames

def menu_km_export_to_excel():

    km_var = km_variable.get()
    filename_main = "{}\jogos_excel.xlsx".format(os.getcwd())
    wb_main = load_workbook(filename_main)
    wb_main['Tabela de prémios']['B14'].value = km_var

    wb_main.save(filename_main)
    time.sleep(0.1)
        
def changeState():
    print(n_arbitro_var.get())
    if n_arbitro_var:
        R1.config(state="normal")
        R2.config(state="normal")
    else:
        pass 
    
#------------MAIN WINDOW CODE------------:

#variáveis:
km_variable = tk.StringVar()
km_variable.set('0.35')
km_quantity = tk.StringVar()
result_km=tk.StringVar()
deslocacao_A = tk.StringVar()
deslocacao_B = tk.StringVar()

n_arbitro_var = IntVar()
selected_option = IntVar()

Label(text='Coloque o seu Nº de Árbitro:').place(x=100,y=10)
tk.Entry(root,textvariable=n_arbitro_var,width=6).place(x=270,y=10)

button_n_arbitro = tk.Button(root,text='OK',command=lambda:[import_excel_doc(), changeState()])
button_n_arbitro.place(x=312,y=6)

options = [
    "0.34", "0.35","0.36","0.37",
    "0.38","0.39","0.40","0.41",
    "0.42","0.43","0.44","0.45"
]

dropdown_text=Label(text='Valor ao KM:').place(x=100,y=50)
menu = tk.OptionMenu(root, km_variable, *options, command=menu_km_export_to_excel())
menu.place(x=175,y=45)

link = Label(root, text='Google Maps',fg='blue',cursor='hand2')
link.place(x=5,y=0)
link.bind("<Button-1>", lambda e: url_link("https://www.google.pt/maps/@38.9954378,-9.1411938,10z?hl=pt-PT"))

#---------Select Values---------
R1 = Radiobutton(root, text="Árbitro Principal", variable=selected_option, value=1,
                 state="disabled",command=lambda:[arbitro_principal(),add_entry_principal()])
R1.place(x=350,y=10)
R2 = Radiobutton(root, text="Todos os Jogos", variable=selected_option, value=2,
                 state="disabled",command=lambda:[Todos_os_Jogos(),add_entry_todos()])
R2.place(x=350,y=35)

button_finish = Button(root,text='Acabei',command=lambda:[lista_of_values(),prepare_to_excel(),export_info_to_excel()]).place(x=590,y=10)



root.resizable(False,False)
root.mainloop()

print(':)')